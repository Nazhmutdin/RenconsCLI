from pathlib import Path
import os
import typing as t

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from src.utils.excel_service_new._types import LoadWorkBookOptions, IterRowsColsOptions


class ExcelService:

    def __init__(self, ws: Worksheet) -> None: ...


    def load_file(self, path: str | Path, **kwargs: t.Unpack[LoadWorkBookOptions]) -> Workbook | t.NoReturn:
        if not os.path.exists(path):
            raise FileNotFoundError(path)

        return load_workbook(path, **kwargs)
    

    @staticmethod
    def append_to_worksheet(data: list[list | dict], ws: Worksheet) -> list[Cell]:

        for row in data:
            ws.append(row)


    def create_workbook(write_only: bool = False, iso_dates: bool = False) -> Workbook:
        return Workbook(write_only, iso_dates)
    

    @staticmethod
    def read_worksheet_by_row(ws: Worksheet, **kwargs: t.Unpack[IterRowsColsOptions]) -> list[list]:
        return [
            list(row) for row in ws.iter_rows(**kwargs)
        ]
    

    @staticmethod
    def read_worksheet_by_column(ws: Worksheet, **kwargs: t.Unpack[IterRowsColsOptions]) -> list[list]:
        return [
            list(column) for column in ws.iter_cols(**kwargs)
        ]
